import sys
import warnings

from PyQt5.QtWidgets import*
from PyQt5.QtCore import *
from PyQt5 import QtGui, uic
from PyQt5 import uic, QtWidgets
from PROYECTOAGN import Ui_MainWindow
import os

from pyparsing import unicode
import re
import time
import itertools



from numpy import integer


# hilos 

import threading
import time

#impor pdf2img 

import PIL.Image, PIL.ImageTk
import pytesseract
import sys
from pdf2image import convert_from_path

# Red Neuronal ....................................................................

import spacy
import fitz
from heapq import nlargest
from collections import Counter
from collections import OrderedDict
from spacy.lang.es.stop_words import STOP_WORDS
from spacy.matcher import PhraseMatcher
from string import punctuation
import string
import re
from unicodedata import normalize
from dateparser.search import search_dates
from datetime import datetime as dt
from hunspell import Hunspell
import pandas as pd


# Excel ....................................................................

import openpyxl 
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
# mongoDB

from pymongo import MongoClient
import os


class Interfaz(QtWidgets.QMainWindow, Ui_MainWindow):

    
    def __init__(self):

        
        #tiene que crearse un archivo de control
        global control
        control ="limpieza/control_archivos.txt"
        #primero hay que saber quien es el user
        global usuario 
        usuario =(os.environ['USERPROFILE']).split("\\").pop()
        #resultadps en  carpeta ocr
        #esta ruta de resultado es incompleta solo hasta la carpeta ocr sin archivo por que despues se adiciona
        global resultados 
        resultados ="limpieza/ocr/"+usuario+"/OCR_"
        
        

        global stopwords
        global nlp
        

        #filePath2 = 'C:/CO_JEP_SNR_RUPTA_106100704/resultado_excel.xlsx'
        

        stopwords = list(STOP_WORDS)
        nlp = spacy.load('es_core_news_md')


        QtWidgets.QMainWindow.__init__(self)
        Ui_MainWindow.__init__(self)
        self.setupUi(self)
        self.setWindowIcon(QtGui.QIcon('AGNICO-03.png'))
        
        self.boton2.clicked.connect(self.frame1)
        self.boton3.clicked.connect(self.archivos_excel)
        self.boton4.clicked.connect(self.Extraer_metadatos)
        self.boton5.clicked.connect(self.frame2)
        #self.boton6.clicked.connect(self.database)
        self.boton7.clicked.connect(self.archivos_individual)
        self.boton8.clicked.connect(self.procesoparalelo)
        self.boton9.clicked.connect(self.agregar_metadatos)
        #self.boton10.clicked.connect(self.eliminar_metadatos)

        self.marcas = {"DESAGREGACIÓN TERRITORIAL" :["NO APLICA"],
               "MUNICIPIOS" :["NO APLICA"],  
               "ACTOR ARMADO": ["FARC-EP", "FUERZAS PUBLICAS", "OTROS AGENTES DEL ESTADO", "OTROS GAOMIL", "TERCEROS"],
               "HECHO VICTIMIZANTE": ["NO APLICA"],
                    "ENFOQUE DIFERENCIAL": ["capacidades diversas", "identidad racial", "niñez y adolescencia", "orientacion sexual y genero", "personas mayores", "perspectiva indigena", "rol social de la victima"],
                        "ENFOQUE TERRITORIAL": ["comunidad de paz", "consejos comunitarios", "resguardos indigena", "zonas de proteccion ambiental", "zonas de reserva campesina"],
                       "RESULTADO METADATOS": ["NO APLICA"],
                       "TITULO DEL DOCUMENTO": ["NO APLICA"],
                       "FECHA DE PRODUCCION": ["NO APLICA"]
                       
                                             }
        
        self.metadatos1.currentIndexChanged[str].connect(self.llenar_comboBox_modelos)
        self.llenar_comboBox_marcas()

        
    def frame1(self):
        self.frame_2.close()
        self.frame_3.close()
        
        
    def frame2(self):
        self.frame_2.show()
        
    
        
    def archivos_excel(self):
        global filePath2

        filePath2, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Select a file...", filter="archivo.xls(*.xlsx)")
        if filePath2 != "":
            print(filePath2)
    


    def borrar_resultados(self):

        folder = "limpieza/ocr/"+usuario
        for the_file in os.listdir(folder):
            file_path = os.path.join(folder, the_file)
            try:
                if os.path.isfile(file_path):
                    os.unlink(file_path)
                # elif os.path.isdir(file_path): shutil.rmtree(file_path)
            except Exception as e:
                #print(e)
                pass

    def crear_carpeta_usuario(self):
        # Se define el nombre de la carpeta o directorio a crear
        directorio = "limpieza/ocr/"+usuario

        try:
            os.mkdir(directorio)
        except OSError:
            #print("La creación del directorio %s falló" % directorio)
            pass
        else:
            #print("Se ha creado el directorio: %s " % directorio)
            pass

            
    def archivos_individual(self):
        
        
        global filePath1
        
        filePath1, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Select a file...", filter="archivos.pdf(*.pdf)")
       
        if filePath1 != "":

            url = QUrl.fromLocalFile(filePath1)
            print(url.fileName())
            self.progreso.setValue(0)
            self.val2.setText("1")
            self.val1.setText("0")

    
            
    def archivos_Lotes(self, data1, data2, hilo_n ):


        
        Client = MongoClient('localhost')
        db = Client['sistema_inteligente_agn']
        Titulo = db['tb_titulo_documento']
        depatarmatetos = db['tb_departamentos']
        Municipios = db['tb_municipios']
        Actor_Armado = db['tb_actor_armado']
        Responsable = db['tb_responsable']
        Victima = db['tb_victimas']
        Enfoque_Diferencial = db['enfoque_diferencial']
        Enfoque_Territorial = db['tb_enfoque_territorial']
        Fecha_Produccion = db['tb_fecha_produccion']
        Fecha_Inicial = db['tb_fecha_inicial']
        Fecha_Final = db['tb_fecha_final']

        fecha_inicial = ""
        fecha_principal = ""
        fecha_cierre = ""
        fecha_doc = ""
        fecha_encontrada_ini = ""
        fecha_encontrada_cierre = ""
        

        print("Inicio proceso No " + hilo_n)
        
        for K in range(data1,data2):

            #lista departamentos
            lisd = []
            #lista municipios
            lisM = []
            #lista actor armado
            lisA = []
            #lista responsable
            lisR = []
            #lista victimas
            lisV = []
            #lista enfoque diferencial
            lisE = []
            #lista enfoque territorial
            lisT = []
            Nivel_D = 2

            
            file_excel_rows = sheet_1.cell(row = K, column = 4).value
            document_1 = file_excel_rows.replace(chr(92),chr(47))
            #####---julian5:32
            disponible=True
            ### verificar que este disponible
            fcontrol = open(control, "r")
            lineas = fcontrol.readlines()

            fcontrol = open(control, "w")
            fcontrol.writelines(lineas)

            for linea in lineas:
                
                if(linea==document_1+'\n'):
                    print("ocupado")
                    disponible=False
                    break

            if(disponible==True):
                fcontrol.write(document_1 + '\n')
            fcontrol.close()

            
            # si  se puedde abrir lo abre y escribe el resultado en un ocr con el nombre del usuario
            if(disponible==True):
                
                # borro ocr anteriores que ya se debieron de haber utilizado
                
                self.crear_carpeta_usuario()
                self.borrar_resultados()

                #creo el nombre de los archivos ocr
                nombre_archivo=document_1.split("/").pop()

                usuario_resultado=resultados+usuario+"_"+nombre_archivo+".txt"
                fresultados = open(usuario_resultado, "x")
                fresultados = open(usuario_resultado, "w")
        ###---julian
            
                if sheet_1.cell(row = K, column = 18).value == None and sheet_1.cell(row = K, column = 7).value == None and sheet_1.cell(row = K, column = 14).value == None and sheet_1.cell(row = K, column = 15).value == None and sheet_1.cell(row = K, column = 16).value == None  and sheet_1.cell(row = K, column = 6).value == None :
                    
                    print(document_1) 
                    document = fitz.open(document_1)
                    url = QUrl.fromLocalFile(document_1)
                    
                    pages = convert_from_path(document_1,size=2500)
                    contador = 1
                    
                    for page in pages:
                
                        filename ="page_"+ str(contador)+"hilo"+hilo_n+".jpg"
                        page.save(filename,'JPEG')
                        contador+=1

                    fileimt = contador-1

                    #f = open("C:/Users/miguel.salas/Desktop/OCR"+hilo_n+".txt","w")
                    f = open("C:/sistema_metadatos/ocr/OCR.txt"+hilo_n+".txt","w")
                
                    for L in range(1,fileimt+1):

                        filename = "page_"+str(L)+"hilo"+hilo_n+".jpg"
                        text = str(((pytesseract.image_to_string(PIL.Image.open(filename)))))
                        text = text.replace('\n', ' ')## se quito el - y se agreg'o espacio  
                        text = re.sub(r'[^a-zA-Z0-9 $¿?.,:;|/|-]+', '', text)
                        if ((len(''.join(ch for ch, _ in itertools.groupby(text)))) > 1):
                            fresultados.write(text + '\n')
                            f.write(text)

                    f.close()
                    fresultados.close()
                        #########'''''''''''''''''''''''''''''------------------------------------------------


                    f = open("C:/sistema_metadatos/ocr/OCR.txt"+hilo_n+".txt","r")
                    text = f.read()
                    f.close()
        
                    doc = nlp(text)
                    tokens = [token.text.replace('\n','') for token in doc]

                    try:

            
                # Dar formato metadato fecha de creaciion del documento...............
                        fec = (document.metadata['creationDate'])
                        fecha_doc = dt.strptime(fec[2:6]+"-"+fec[6:8]+"-"+fec[8:10],'%Y-%m-%d')

                # Fecha Produccion....................................................
                        text_fecha = text[0:500]
                        fecha_1 = search_dates(text_fecha, languages=['es'])

                        fecha_principal = 'No Registra'

                        validacion_fechas = []
                        if fecha_1 != None:
                            for n in range(len(fecha_1)):
                                if len(str(pd.DataFrame(fecha_1).iloc[n,0])) > 7:
                                    validacion_fechas.append(pd.DataFrame(fecha_1).iloc[n,1])
                        validacion_fechas = pd.DataFrame(validacion_fechas)        
                        if len(validacion_fechas) >= 1:
                            validacion_fechas = validacion_fechas[validacion_fechas[0].isin(pd.date_range('1950-1-1', fecha_doc))]
                            if len(validacion_fechas) >= 1:
                                fecha_principal = pd.Timestamp.date(validacion_fechas[0][0])
                    #print(fecha_principal)
                
                    except: 
                        print("An exception occurred error en fecha de produccion")
                        fecha_inicial = "error"
                # fecha inicio proceso ..............................................................
                    # Activacion PhraseMatcher fecha inicio
                    matcher_ini = PhraseMatcher(nlp.vocab)
                        
                    # Preparacion tag desde la BD fecha inicio
                    terms_f_ini_proceso = list(pd.DataFrame(Fecha_Inicial.find())['VC_Tag'])
                                
                    # terminos a tipo doc inicio proceso
                    patterns_ini = [nlp.make_doc(tags) for tags in terms_f_ini_proceso]
                    matcher_ini.add("List_f_ini", patterns_ini)

                    # Preparacion del texto para fechas del proceso
                    if text.isupper: text_lower = text.lower()
                    text_lower = re.sub( r"([^n\u0300-\u036f]|n(?!\u0303(?![\u0300-\u036f])))[\u0300-\u036f]+", r"\1", normalize( "NFD", text_lower), 0, re.I)
                    doc_log_patterns = nlp(text_lower)

                    # Busqueda de textos que conincidan fecha inicio
                    matches_ini = matcher_ini(doc_log_patterns)
                        
                    texto_fechas_ini =[]
                    try:
                        for match_id, start, end in matches_ini:
                            span = doc_log_patterns[start:end+15]
                            texto_fechas_ini.append(str(span))
                
                    except: print(" error en fecha de inicio - 1") 


                    # Extraccion fechas
                    lista_f_e_i = []
                
                    try:
                        for fechas in texto_fechas_ini:            
                            fecha_encontrada_ini = search_dates(fechas, languages=['es'])
                            if fecha_encontrada_ini != None:
                                for n in range(len(fecha_encontrada_ini)):
                                    if len(str(pd.DataFrame(fecha_encontrada_ini).iloc[n,0])) > 7:
                                        lista_f_e_i.append(pd.DataFrame(fecha_encontrada_ini).iloc[n,1])
                    # Filtro fechas
                        fecha_inicial = 'No registra'
                        lista_f_e_i = pd.DataFrame(lista_f_e_i)
                        if len(lista_f_e_i) >= 1:
                            lista_f_e_i = lista_f_e_i[lista_f_e_i[0].isin(pd.date_range('1950-1-1', fecha_doc))]
                            if len(lista_f_e_i) >= 1:
                                fecha_inicial = pd.Timestamp.date(max(lista_f_e_i[0]))
                    except(): print("error en fecha de inicio - 2")
        # fecha cierre proceso ..................................................................................

                    # Activacion PhraseMatcher fecha cierre
                    matcher_cierre = PhraseMatcher(nlp.vocab)
                        
                    # Preparacion tag desde la BD fecha cierre
                    terms_f_cierre = list(pd.DataFrame(Fecha_Final.find())['VC_Tag'])
                                    
                    # terminos a tipo doc cierre proceso
                    patterns_cierre = [nlp.make_doc(tags) for tags in terms_f_cierre]
                    matcher_cierre.add("List_f_cierre", patterns_cierre)

                    # Busqueda de textos que conincidan fecha cierre
                    matches_cierre = matcher_cierre(doc_log_patterns)
                        
                    texto_fechas_cierre =[]

                    try:
                        for match_id, start, end in matches_cierre:
                            span = doc_log_patterns[start:end+15]
                            texto_fechas_cierre.append(str(span))
                    
                
                        lista_f_e_c = []
                        for fechas in texto_fechas_cierre:            
                            fecha_encontrada_cierre = search_dates(fechas, languages=['es'])
                        if fecha_encontrada_cierre != None:
                            for n in range(len(fecha_encontrada_cierre)):
                                if len(str(pd.DataFrame(fecha_encontrada_cierre).iloc[n,0])) > 7:
                                    lista_f_e_c.append(pd.DataFrame(fecha_encontrada_cierre).iloc[n,1])
                    
                    except():print("error en fecha de cierre - 1")
                    fecha_cierre = 'No registra'
                    
                    try:
                        lista_f_e_c = pd.DataFrame(lista_f_e_c)
                        if len(lista_f_e_c) >= 1:
                            lista_f_e_c = lista_f_e_c[lista_f_e_c[0].isin(pd.date_range('1950-1-1', fecha_doc))]
                            if len(lista_f_e_c) >= 1:
                                fecha_cierre = pd.Timestamp.date(min(lista_f_e_c[0]))
                    except(): print("error en fecha de cierre - 2")
            # Titulo del documento .....................................................................................................
                    # Activacion PhraseMatcher
                    matcher_titulo = PhraseMatcher(nlp.vocab)

                    # Preparacion tag desde la BD
                    terms_titulo = list(pd.DataFrame(Titulo.find())['VC_Tag'])

                    # terminos a tipo doc titulo documento
                    patterns_titulo = [nlp.make_doc(tags) for tags in terms_titulo]
                    matcher_titulo.add("List_titulo", patterns_titulo)

                    # Fracmeto de texto para buscar titulo
                    lineas = re.split(r'\n',text_lower[0:500])

                    # Extraccion de coincidencias titulo
                    titulo = []
                    for ti in lineas:
                        tok =nlp(ti)
                        matches_titulo = matcher_titulo(tok)
                        list_titulo = []
                        for match_id, start, end in matches_titulo:
                            span = tok[start-50:end+50]
                        if len(matches_titulo) >= 1:
                            titulo.append([ti,span])
                    # Extraccion cadena mas larga del titulo
                    if len(titulo) >= 1:
                        titulo_doc =[]
                        for tid in titulo:
                            titulo_doc.append([len(tid[0]),tid])

                        if len(titulo_doc) >=1: 
                            titulo = str(max(titulo_doc)[1][1])
                        #print('Titulo:',titulo)
                    if len(titulo) < 1:
                        titulo ='Titulo no encontrado'

            # filtro uno .................................................................................................................


                    for palabra in tokens:

                        partes_frase = re.split(r'\n',palabra)  

                        for partes in partes_frase:
            
                            if partes.isupper: partes = partes.lower() 
                            partes = re.sub( r"([^n\u0300-\u036f]|n(?!\u0303(?![\u0300-\u036f])))[\u0300-\u036f]+", r"\1", normalize( "NFD", partes), 0, re.I)
    
            # Departamentos ........................................................................................

                            doc_1 = depatarmatetos.find_one({"VC_Departamento":partes})
                
                            if doc_1 != None:

                    #lisd.append(doc_1['VC_Departamento'])
                                lisd.append(doc_1['IN_Codigo_Departamento'])

                                if Nivel_D == 2: Nivel_D = 4

                # Municipios ........................................................................................
                
        
                            doc_1 = Municipios.find_one({"VC_Municipio":partes})
                        
                            if doc_1 != None:
                                
                #lisM.append(doc_1['VC_Municipio'])
                                lisM.append(doc_1['_id'])

                                if Nivel_D == 2 or Nivel_D == 4: Nivel_D = 5
            

                    ### ACTOR ARMADO......................................................................................
                            doc_1 = Actor_Armado.find_one({"VC_Tag":partes})     
        
                            if doc_1 != None:
            
                    #lisA.append(doc_1['VC_Tag'])
                                lisA.append(doc_1['_id'])
                ### RESPONSABLE ......................................................................................
                            doc_1 = Responsable.find_one({"VC_Tag":partes})     
        
                            if doc_1 != None:
            
                    #lisA.append(doc_1['VC_Tag'])
                                lisR.append(doc_1['_id'])

                #VICTIMA.................................................................................................................... 
                            doc_1 = Victima.find_one({"VC_Tag":partes})       
            
                            if doc_1 != None:
            
                    #lisA.append(doc_1['VC_Tag'])
                                lisV.append(doc_1['_id'])

                ### ENFOQUE DIFERENCIAL.............................................................................................. 
                    
                            doc_1 = Enfoque_Diferencial.find_one({"VC_Tag":partes})       
                
                            if doc_1 != None:
                
                    #lisA.append(doc_1['VC_Tag'])
                                lisE.append(doc_1['_id'])
            
                ### ENFOQUE TERRITORIAL      
        
                            doc_1 = Enfoque_Territorial.find_one({"VC_Tag":partes})       
    
                            if doc_1 != None:
            
                    #lisA.append(doc_1['VC_Tag'])
                                lisT.append(doc_1['_id']) 

            
                    for palabra in doc.ents:
        
                        partes_frase = re.split(r'\n',palabra.text)  

                        for partes in partes_frase:
            
                            if partes.isupper: partes = partes.lower() 
                            partes = re.sub( r"([^n\u0300-\u036f]|n(?!\u0303(?![\u0300-\u036f])))[\u0300-\u036f]+", r"\1", normalize( "NFD", partes), 0, re.I)
            

                ### ACTOR ARMADO......................................................................................
                            doc_1 = Actor_Armado.find_one({"VC_Tag":partes})     
        
                            if doc_1 != None:
            
                    #lisA.append(doc_1['VC_Tag'])
                                lisA.append(doc_1['_id'])

                ### RESPONSABLE ......................................................................................
                            doc_1 = Responsable.find_one({"VC_Tag":partes})     
        
                            if doc_1 != None:
                    #lisR.append(doc_1['VC_Tag'])
                                lisR.append(doc_1['_id'])            

        
                #VICTIMA...................................................................................................................
                            doc_1 = Victima.find_one({"VC_Tag":partes})       
            
                            if doc_1 != None:
                
                    #lisA.append(doc_1['VC_Tag'])
                                lisV.append(doc_1['_id'])

                ### ENFOQUE DIFERENCIAL.............................................................................................. 
                    
                            doc_1 = Enfoque_Diferencial.find_one({"VC_Tag":partes})       
                
                            if doc_1 != None:
            
        
                                lisE.append(doc_1['_id'])
            
                ### ENFOQUE TERRITORIAL      
        
                            doc_1 = Enfoque_Territorial.find_one({"VC_Tag":partes})       
    
                            if doc_1 != None:  
            
                                lisT.append(doc_1['_id']) 
            

                            if palabra.label_ in ['LOC']: 

                                if partes == "veredal" or "vereda": Nivel_D = 6
                
                # Departamentos ........................................................................................

                            doc_1 = depatarmatetos.find_one({"VC_Departamento":partes})
                    
                            if doc_1 != None:

                    #lisd.append(doc_1['VC_Departamento'])
                                lisd.append(doc_1['IN_Codigo_Departamento'])
                                if Nivel_D == 2: Nivel_D = 4

                # Municipios ........................................................................................
                    

                            doc_1 = Municipios.find_one({"VC_Municipio":partes})
                        
                            if doc_1 != None:
                                
                        #lisM.append(doc_1['VC_Municipio'])
                                lisM.append(doc_1['_id'])

                                if Nivel_D == 2 or Nivel_D == 4: Nivel_D = 5


                    if bool(lisd) == False and bool(lisM) == False: Nivel_D = 8
                    if bool(lisA) == False: lisA.append(7)
                    if bool(lisR) == False: lisR.append(0)
                    if bool(lisV) == False: lisV.append(0) 
                    if bool(lisE) == False: lisE.append(9)
                    if bool(lisT) == False: lisT.append(7)

        # verificamos frecuencia .....................................

                    word_frequencies = {}
                    for word in doc:
                        if word.text.lower() not in stopwords:
                            if word.text.lower() not in punctuation:
                                if word.text not in word_frequencies.keys():
                                    word_frequencies[word.text] = 1
                            else:
                                word_frequencies[word.text] =+ 1


                    max_frequency = max(word_frequencies.values())

        # acotamos ......................................

                    for word in word_frequencies.keys():
                        word_frequencies[word] = word_frequencies[word]/max_frequency

        # obtiene oraciones ...........................................

                    sentence_tokens = [sent for sent in doc.sents]

                # ordenamos oraciones ........................................ 

                    sentence_scores = {}
                    for sent in sentence_tokens:
                        for word in sent:
                            if word.text.lower() in word_frequencies.keys():
                                if sent not in sentence_scores.keys():
                                    sentence_scores[sent] = word_frequencies[word.text.lower()]
                                else:
                                    sentence_scores[sent] += word_frequencies[word.text.lower()] 

                # obtiene 30% del resumen 

                    select_length = int(len(sentence_tokens)*0.30)
                    summary = nlargest(select_length, sentence_scores,key=sentence_scores.get)
                    final_summary = [word.text for word in summary]
                    summary = ''.join(final_summary)
                    summary = ILLEGAL_CHARACTERS_RE.sub(r'', summary)
                
                    resultD = " "


                    if sheet_1.cell(row = K, column = 5).value == url.fileName():

                        
                        sheet_1.cell(row = K, column = 8).value =  str(Nivel_D) 
                    


                        lisd = set(lisd)
                        lisd = list(lisd)

                        for j in range(0,len(lisd)):
                
                                if len(lisd) == 1: resultD =  str(int(lisd[j])) + resultD
                                if lisd != " " and len(lisd) > 1: resultD = str(int(lisd[j])) + '|' + resultD
                                

                        sheet_1.cell(row = K, column = 9).value =  resultD
                        resultD = " "

                        lisM = set(lisM)
                        lisM = list(lisM)

                        for j in range(0,len(lisM)):
                                    
                                if len(lisM) == 1: resultD =  str(int(lisM[j])) + resultD
                                if lisM != " " and len(lisM) > 1: resultD = str(int(lisM[j])) + '|' + resultD
                                

                        sheet_1.cell(row = K, column = 10).value =  resultD
                        resultD = " "
                                
                        lisA = set(lisA)
                        lisA = list(lisA)
                        
                        for j in range(0,len(lisA)):

                                    
                            if len(lisA) == 1: resultD =  str(int(lisA[j])) + resultD
                            if lisA != " " and len(lisA) > 1: resultD = str(int(lisA[j])) + '|' + resultD
                                
                            
                        sheet_1.cell(row = K, column = 11).value = resultD
                        resultD = " "

                        for j in set(lisR):  
    
                            if lisR != " ": sheet_1.cell(row = K, column = 12).value = j

                        for j in set(lisV):

                            if lisV != " ": sheet_1.cell(row = K, column = 13).value =  j

                        for j in range(0,len(set(lisE))):
                            
                            if len(lisE) == 1: resultD =  str(int(lisE[j])) + resultD
                            if lisE != " " and len(lisE) > 1: resultD = str(int(lisE[j])) + '|' + resultD
                                
                                
                        sheet_1.cell(row = K, column = 17).value =  resultD
                        resultD = " "

                        lisT = set(lisT)
                        lisT = list(lisT)

                        for j in range(0,len(lisT)):

                            if len(lisT) == 1: resultD =  str(int(lisT[j])) + resultD
                            if lisT != " " and len(lisT) > 1: resultD = str(int(lisT[j])) + '|' + resultD
                        
                                
                        sheet_1.cell(row = K, column = 18).value =  resultD
                        sheet_1.cell(row = K, column = 7).value =   summary
                        sheet_1.cell(row = K, column = 14).value =   fecha_principal
                        sheet_1.cell(row = K, column = 15).value =  fecha_inicial
                        sheet_1.cell(row = K, column = 16).value =  fecha_cierre  
                        sheet_1.cell(row = K, column = 6).value =  titulo 

                    worbook.save(filePath2)
            
                

                

                # parte final para liberar archivo
                #quito el archivo que se estaba utilizando DEL ARCHIVO DE CONTROL  dejando las lineas actuales
                fcontrol = open(control, "r")
                lineas=fcontrol.readlines()

                fcontrol = open(control, "w")
                for linea in lineas:
                    #print(linea +"  "+ document_1+'\n' )
                    if (linea != document_1+'\n' ):
                        fcontrol.write(linea)
                    else:
                        pass
                fcontrol.close()
            else:
                print("archivo ocupado")
                    #### fin metodo
        print("fin de Proceso No "+ hilo_n)
        # self.ui.progreso.setValue(porcentaje_1)
    
    def procesoparalelo(self):

        global ContDocumento
        global tokens
        global url
        global worbook
        global sheet_1
        global datos_inicial
        global datos_lotes 

        ContDocumento = 0
        porcentaje_1 = 0

        worbook = openpyxl.load_workbook(filePath2,data_only=True)
        sheet_1 = worbook.active

        dato_a = sheet_1.max_row - 1
       
        self.val2.setText(str(dato_a))
        self.val1.setText(str(ContDocumento))


        
        self.progreso.setValue(0)
        
       

        self.archivos_Lotes(2,4,"1")
        

         #t1 = threading.Thread(name ="ex1",target= self.archivos_Lotes, args =([data_i,data_f,"1"] ))
         #t2 = threading.Thread(name ="ex2",target= self.archivos_Lotes, args =([data_i2,data_f2,"2"]))
         #t3 = threading.Thread(name ="ex3",target= self.archivos_Lotes, args =([data_i3,data_f3,"3"]))
         #t4 = threading.Thread(name ="ex4",target= self.archivos_Lotes, args =([data_i4,data_f4,"4"]))
         #t5 = threading.Thread(name ="ex5",target= self.archivos_Lotes, args =([data_i5,data_f5,"5"]))
         #t6 = threading.Thread(name ="ex6",target= self.archivos_Lotes, args =([data_i6,data_f6,"6"]))

        #t1.start() 
         #t2.start()
         #t3.start()
         #t4.start()
         #t5.start()
         #t6.start()

   


        
    def Extraer_metadatos(self):

        
        disponible=True
        ### verificar que este disponible
        fcontrol = open(control, "r")
        lineas = fcontrol.readlines()

        fcontrol = open(control, "w")
        fcontrol.writelines(lineas)

        for linea in lineas:
            
            if(linea==filePath1+'\n'):
                print("ocupado")
                disponible=False
                break

        if(disponible==True):
            fcontrol.write(filePath1 + '\n')
        fcontrol.close()

        
        # si  se puedde abrir lo abre y escribe el resultado en un ocr con el nombre del usuario
        if(disponible==True):
            
            # borro ocr anteriores que ya se debieron de haber utilizado
            
            self.crear_carpeta_usuario()
            self.borrar_resultados()

            #creo el nombre de los archivos ocr
            nombre_archivo=filePath1.split("/").pop()

            usuario_resultado=resultados+usuario+"_"+nombre_archivo+".txt"
            fresultados = open(usuario_resultado, "x")
            fresultados = open(usuario_resultado, "w")

        
            Client = MongoClient('localhost')
            db = Client['sistema_inteligente_agn']
            Titulo = db['tb_titulo_documento']
            depatarmatetos = db['tb_departamentos']
            Municipios = db['tb_municipios']
            Actor_Armado = db['tb_actor_armado']
            Responsable = db['tb_responsable']
            Victima = db['tb_victimas']
            Enfoque_Diferencial = db['enfoque_diferencial']
            Enfoque_Territorial = db['tb_enfoque_territorial']
            Fecha_Produccion = db['tb_fecha_produccion']
            Fecha_Inicial = db['tb_fecha_inicial']
            Fecha_Final = db['tb_fecha_final']

            """dic = Hunspell("es_ANY")

            for doc_add in depatarmatetos.find():
                dic.add(doc_add["VC_Departamento"])

            for doc_add in Municipios.find():
                dic.add(doc_add["VC_Municipio"])"""

        
            global tokens
            global url, FILE_PATH

            porcentaje = 100

            FILE_PATH = filePath2
            self.val2.setText("1")
            self.val1.setText("0")
        

            worbook = openpyxl.load_workbook(FILE_PATH,data_only=True)
            sheet_1 = worbook.active
            self.progreso.setValue(3)
    
            lisd = []
            lisM = []
            lisA = []
            lisR = []
            lisV = []
            lisE = []
            lisT = []
            Nivel_D = 2
            
            
            document = fitz.open(filePath1)
            url = QUrl.fromLocalFile(filePath1)
            
            pages = convert_from_path(filePath1,size=2500)
            contador = 1 
            
            for page in pages:
                
                filename ="page_"+ str(contador)+".jpg"
                page.save(filename,'JPEG')
                contador+=1

            fileimt = contador-1

            f = open("C:/sistema_metadatos/ocr/OCR.txt","w")

            for i in range(1,fileimt+1):

                filename = "page_"+str(i)+".jpg"
                text = str(((pytesseract.image_to_string(PIL.Image.open(filename)))))
                text = text.replace('-\n', '')  

                text = re.sub(r'[^a-zA-Z0-9 $¿?.,:;|/|-]+', '', text)
                if ((len(''.join(ch for ch, _ in itertools.groupby(text)))) > 1):
                    fresultados.write(text + '\n')
                    f.write(text)

            f.close()
            fresultados.close()

            f = open('C:/sistema_metadatos/ocr/OCR.txt','r')
            text = f.read()
            f.close()
    
            doc = nlp(text)
            tokens = [token.text.replace('\n','') for token in doc]


            self.progreso.setValue(int(porcentaje*0.1))

    # Dar formato metadato fecha de creaciion del documento...........
            fec = (document.metadata['creationDate'])
            #fecha_doc = dt.strptime(fec[2:6]+"-"+fec[6:8]+"-"+fec[8:10],'%Y-%m-%d')
            fecha_doc = dt.strptime('2021-01-01','%Y-%m-%d')
            #print(fec)

    # Fecha Produccion...................................
            text_fecha = text[0:500]
            fecha_1 = search_dates(text_fecha, languages=['es'])

            fecha_principal = 'No Registra'

            validacion_fechas = []
            if fecha_1 != None:
                for n in range(len(fecha_1)):
                    if len(str(pd.DataFrame(fecha_1).iloc[n,0])) > 7:
                        validacion_fechas.append(pd.DataFrame(fecha_1).iloc[n,1])
            validacion_fechas = pd.DataFrame(validacion_fechas)        
            if len(validacion_fechas) >= 1:
                validacion_fechas = validacion_fechas[validacion_fechas[0].isin(pd.date_range('1950-1-1', fecha_doc))]
                if len(validacion_fechas) >= 1:
                    fecha_principal = pd.Timestamp.date(validacion_fechas[0][0])
            #print(fecha_principal)
        

    # fecha inicio proceso ..............................................................
            # Activacion PhraseMatcher fecha inicio
            matcher_ini = PhraseMatcher(nlp.vocab)
                    
            # Preparacion tag desde la BD fecha inicio
            terms_f_ini_proceso = list(pd.DataFrame(Fecha_Inicial.find())['VC_Tag'])
                            
            # terminos a tipo doc inicio proceso
            patterns_ini = [nlp.make_doc(tags) for tags in terms_f_ini_proceso]
            matcher_ini.add("List_f_ini", patterns_ini)

            # Preparacion del texto para fechas del proceso
            if text.isupper: text_lower = text.lower()
            text_lower = re.sub( r"([^n\u0300-\u036f]|n(?!\u0303(?![\u0300-\u036f])))[\u0300-\u036f]+", r"\1", normalize( "NFD", text_lower), 0, re.I)
            doc_log_patterns = nlp(text_lower)

            # Busqueda de textos que conincidan fecha inicio
            matches_ini = matcher_ini(doc_log_patterns)
                
            texto_fechas_ini =[]
            for match_id, start, end in matches_ini:
                span = doc_log_patterns[start:end+15]
                texto_fechas_ini.append(str(span))


            # Extraccion fechas
            lista_f_e_i = []
            for fechas in texto_fechas_ini:            
                fecha_encontrada_ini = search_dates(fechas, languages=['es'])
                if fecha_encontrada_ini != None:
                    for n in range(len(fecha_encontrada_ini)):
                        if len(str(pd.DataFrame(fecha_encontrada_ini).iloc[n,0])) > 7:
                            lista_f_e_i.append(pd.DataFrame(fecha_encontrada_ini).iloc[n,1])
            # Filtro fechas
            fecha_inicial = 'No registra'
            lista_f_e_i = pd.DataFrame(lista_f_e_i)
            if len(lista_f_e_i) >= 1:
                lista_f_e_i = lista_f_e_i[lista_f_e_i[0].isin(pd.date_range('1950-1-1', fecha_doc))]
                if len(lista_f_e_i) >= 1:
                    fecha_inicial = pd.Timestamp.date(max(lista_f_e_i[0]))
            


    # fecha cierre proceso ..................................................................................

            # Activacion PhraseMatcher fecha cierre
            matcher_cierre = PhraseMatcher(nlp.vocab)
                    
            # Preparacion tag desde la BD fecha cierre
            terms_f_cierre = list(pd.DataFrame(Fecha_Produccion.find())['VC_Tag'])
                            
            # terminos a tipo doc cierre proceso
            patterns_cierre = [nlp.make_doc(tags) for tags in terms_f_cierre]
            matcher_cierre.add("List_f_cierre", patterns_cierre)

            # Busqueda de textos que conincidan fecha cierre
            matches_cierre = matcher_cierre(doc_log_patterns)
                
            texto_fechas_cierre =[]
            for match_id, start, end in matches_cierre:
                span = doc_log_patterns[start:end+15]
                texto_fechas_cierre.append(str(span))
                
            
            lista_f_e_c = []
            for fechas in texto_fechas_cierre:            
                fecha_encontrada_cierre = search_dates(fechas, languages=['es'])
                if fecha_encontrada_cierre != None:
                    for n in range(len(fecha_encontrada_cierre)):
                        if len(str(pd.DataFrame(fecha_encontrada_cierre).iloc[n,0])) > 7:
                            lista_f_e_c.append(pd.DataFrame(fecha_encontrada_cierre).iloc[n,1])
                

            fecha_cierre = 'No registra'

            lista_f_e_c = pd.DataFrame(lista_f_e_c)
            if len(lista_f_e_c) >= 1:
                lista_f_e_c = lista_f_e_c[lista_f_e_c[0].isin(pd.date_range('1950-1-1', fecha_doc))]
                if len(lista_f_e_c) >= 1:
                    fecha_cierre = pd.Timestamp.date(min(lista_f_e_c[0]))
            
    # Titulo del documento .....................................................................................................

            # Activacion PhraseMatcher
            matcher_titulo = PhraseMatcher(nlp.vocab)

            # Preparacion tag desde la BD
            terms_titulo = list(pd.DataFrame(Titulo.find())['VC_Tag'])

            # terminos a tipo doc titulo documento
            patterns_titulo = [nlp.make_doc(tags) for tags in terms_titulo]
            matcher_titulo.add("List_titulo", patterns_titulo)

            # Fracmeto de texto para buscar titulo
            lineas = re.split(r'\n',text_lower[0:500])

            # Extraccion de coincidencias titulo
            titulo = []
            for ti in lineas:
                tok =nlp(ti)
                matches_titulo = matcher_titulo(tok)
                list_titulo = []
                for match_id, start, end in matches_titulo:
                    span = tok[start-50:end+50]
                if len(matches_titulo) >= 1:
                    titulo.append([ti,span])
            # Extraccion cadena mas larga del titulo
            if len(titulo) >= 1:
                titulo_doc =[]
                for tid in titulo:
                    titulo_doc.append([len(tid[0]),tid])
                if len(titulo_doc) >=1: 
                    titulo = str(max(titulo_doc)[1][1])
                #print('Titulo:',titulo)
            if len(titulo) < 1:
                titulo ='Titulo no encontrado'

    # filtro uno .................................................................................................................


            for palabra in tokens:

                partes_frase = re.split(r'\n',palabra)  

                for partes in partes_frase:
            
                    if partes.isupper: partes = partes.lower() 
                    partes = re.sub( r"([^n\u0300-\u036f]|n(?!\u0303(?![\u0300-\u036f])))[\u0300-\u036f]+", r"\1", normalize( "NFD", partes), 0, re.I)
    
                # Departamentos ........................................................................................

                    doc_1 = depatarmatetos.find_one({"VC_Departamento":partes})
                
                    if doc_1 != None:

                    #lisd.append(doc_1['VC_Departamento'])
                        lisd.append(doc_1['IN_Codigo_Departamento'])

                        if Nivel_D == 2: Nivel_D = 4

                # Municipios ........................................................................................
                
        
                    doc_1 = Municipios.find_one({"VC_Municipio":partes})
                    
                    if doc_1 != None:
                            
                #lisM.append(doc_1['VC_Municipio'])
                        lisM.append(doc_1['_id'])

                        if Nivel_D == 2 or Nivel_D == 4: Nivel_D = 5


                    ### ACTOR ARMADO......................................................................................
                    doc_1 = Actor_Armado.find_one({"VC_Tag":partes})     
        
                    if doc_1 != None:
            
                #lisA.append(doc_1['VC_Tag'])
                        lisA.append(doc_1['_id'])


            ### RESPONSABLE ......................................................................................
                    doc_1 = Responsable.find_one({"VC_Tag":partes})     
        
                    if doc_1 != None:
            
                #lisA.append(doc_1['VC_Tag'])
                        lisR.append(doc_1['_id'])

        
                #VICTIMA.................................................................................................................... 
                    doc_1 = Victima.find_one({"VC_Tag":partes})       
        
                    if doc_1 != None:
            
                #lisA.append(doc_1['VC_Tag'])
                        lisV.append(doc_1['_id'])

                ### ENFOQUE DIFERENCIAL.............................................................................................. 
                    
                    doc_1 = Enfoque_Diferencial.find_one({"VC_Tag":partes})       
            
                    if doc_1 != None:
            
                #lisA.append(doc_1['VC_Tag'])
                        lisE.append(doc_1['_id'])
            
                ### ENFOQUE TERRITORIAL      
        
                    doc_1 = Enfoque_Territorial.find_one({"VC_Tag":partes})       
    
                    if doc_1 != None:
            
                    #lisA.append(doc_1['VC_Tag'])
                        lisT.append(doc_1['_id']) 

            self.progreso.setValue(int(porcentaje*0.5)) 

            for palabra in doc.ents:
        
                partes_frase = re.split(r'\n',palabra.text)  

                for partes in partes_frase:
            
                    if partes.isupper: partes = partes.lower() 
                    partes = re.sub( r"([^n\u0300-\u036f]|n(?!\u0303(?![\u0300-\u036f])))[\u0300-\u036f]+", r"\1", normalize( "NFD", partes), 0, re.I)
            

        ### ACTOR ARMADO......................................................................................
                    doc_1 = Actor_Armado.find_one({"VC_Tag":partes})     
        
                    if doc_1 != None:
            
            #lisA.append(doc_1['VC_Tag'])
                        lisA.append(doc_1['_id'])

            ### RESPONSABLE ......................................................................................
                    doc_1 = Responsable.find_one({"VC_Tag":partes})     
        
                    if doc_1 != None:
            
            #lisR.append(doc_1['VC_Tag'])
                        lisR.append(doc_1['_id'])            

        
        #VICTIMA...................................................................................................................
                    doc_1 = Victima.find_one({"VC_Tag":partes})       
        
                    if doc_1 != None:
            
            #lisA.append(doc_1['VC_Tag'])
                        lisV.append(doc_1['_id'])

            ### ENFOQUE DIFERENCIAL.............................................................................................. 
                    
                    doc_1 = Enfoque_Diferencial.find_one({"VC_Tag":partes})       
            
                    if doc_1 != None:
            
        
                        lisE.append(doc_1['_id'])
            
        ### ENFOQUE TERRITORIAL      
        
                    doc_1 = Enfoque_Territorial.find_one({"VC_Tag":partes})       
    
                    if doc_1 != None:  
            
                        lisT.append(doc_1['_id']) 
            

    
                    if palabra.label_ in ['LOC']: 

                        if partes == "veredal" or "vereda": Nivel_D = 6
                
                # Departamentos ........................................................................................

                    doc_1 = depatarmatetos.find_one({"VC_Departamento":partes})
                
                    if doc_1 != None:

                    #lisd.append(doc_1['VC_Departamento'])
                        lisd.append(doc_1['IN_Codigo_Departamento'])
                        if Nivel_D == 2: Nivel_D = 4

                # Municipios ........................................................................................
                
        
                    doc_1 = Municipios.find_one({"VC_Municipio":partes})
                    
                    if doc_1 != None:
                            
                    #lisM.append(doc_1['VC_Municipio'])
                        lisM.append(doc_1['_id'])

                        if Nivel_D == 2 or Nivel_D == 4: Nivel_D = 5

                    


            if bool(lisd) == False and bool(lisM) == False: Nivel_D = 8
            if bool(lisA) == False: lisA.append(7)
            if bool(lisR) == False: lisR.append(0)
            if bool(lisV) == False: lisV.append(0) 
            if bool(lisE) == False: lisE.append(9)
            if bool(lisT) == False: lisT.append(7)

                    
    # verificamos frecuencia .....................................

            word_frequencies = {}
            for word in doc:
                if word.text.lower() not in stopwords:
                    if word.text.lower() not in punctuation:
                        if word.text not in word_frequencies.keys():
                            word_frequencies[word.text] = 1
                    else:
                        word_frequencies[word.text] =+ 1


            max_frequency = max(word_frequencies.values())

    # acotamos ......................................

            for word in word_frequencies.keys():
                word_frequencies[word] = word_frequencies[word]/max_frequency

    # obtiene oraciones ...........................................

            sentence_tokens = [sent for sent in doc.sents]

    # ordenamos oraciones ........................................ 

            sentence_scores = {}
            for sent in sentence_tokens:
                for word in sent:
                    if word.text.lower() in word_frequencies.keys():
                        if sent not in sentence_scores.keys():
                            sentence_scores[sent] = word_frequencies[word.text.lower()]
                        else:
                            sentence_scores[sent] += word_frequencies[word.text.lower()] 

    # obtiene 30% del resumen 

            select_length = int(len(sentence_tokens)*0.90)
            summary = nlargest(select_length, sentence_scores,key=sentence_scores.get)
            final_summary = [word.text for word in summary]
            summary = ''.join(final_summary)
            summary = ILLEGAL_CHARACTERS_RE.sub(r'', summary)

            resumen_final = 'El documento '+ url.fileName() +' fue producido en la fecha'+ document.metadata['creationDate'] +'el documento contiene '+ str(document.pageCount) +'paginas, donde se encontraron '+ str(len(doc)) +' Palabras. :\n  El resumen es el siguiente: :\n'
            resumen_final = resumen_final + summary
            
            resultD = " "
            
            m_row = sheet_1.max_row
            self.progreso.setValue(int(porcentaje*0.8))
            #print(filePath2)
            
            
            for i in range(1, m_row + 1):

                
                    
                    if sheet_1.cell(row = i, column = 5).value == url.fileName():
                        
                        sheet_1.cell(row = i, column = 8).value =  str(Nivel_D) + '|'

                        lisd = set(lisd)
                        lisd = list(lisd)

                        for j in range(0,len(lisd)):
                

                            if lisd != " ": resultD = str(int(lisd[j])) + '|' + resultD

                        sheet_1.cell(row = i, column = 9).value =  resultD
                        resultD = " "

                        lisM = set(lisM)
                        lisM = list(lisM)

                        for j in range(0,len(lisM)):
                            
                            if lisM != " ":  resultD = str(int(lisM[j])) + '|' + resultD

                        sheet_1.cell(row = i, column = 10).value =  resultD
                        resultD = " "

                        
                        lisA = set(lisA)
                        lisA = list(lisA)

                        
                        for j in range(0,len(lisA)):
                            
                            if lisA != " ": resultD =  str(int(lisA[j])) + '|' + resultD
                            
                        sheet_1.cell(row = i, column = 11).value = resultD
                        resultD = " "

                        for j in set(lisR):  
    
                            if lisR != " ": sheet_1.cell(row = i, column = 12).value = j

                        for j in set(lisV):

                            if lisV != " ": sheet_1.cell(row = i, column = 13).value =  j

                        for j in range(0,len(set(lisE))):
                            
                            if lisE != " ": resultD =  str(int(lisE[j])) + '|' + resultD
                                
                        sheet_1.cell(row = i, column = 17).value =  resultD
                        resultD = " "

                        
                        lisT = set(lisT)
                        lisT = list(lisT)

                        for j in range(1,len(lisT)):
                            
                            if lisT != " ": resultD =  str(int(lisT[j])) + '|' + resultD
                        
                        sheet_1.cell(row = i, column = 18).value =  resultD
                        sheet_1.cell(row = i, column = 7).value =   resumen_final
                        sheet_1.cell(row = i, column = 14).value =   fecha_principal
                        sheet_1.cell(row = i, column = 15).value =  fecha_inicial
                        sheet_1.cell(row = i, column = 16).value =  fecha_cierre  
                        sheet_1.cell(row = i, column = 6).value =  titulo 


            worbook.save(FILE_PATH)
            self.progreso.setValue(int(porcentaje))
            self.val2.setText("1")
            self.val1.setText("1")

            # parte final para liberar archivo
            #quito el archivo que se estaba utilizando DEL ARCHIVO DE CONTROL  dejando las lineas actuales
            fcontrol = open(control, "r")
            lineas=fcontrol.readlines()

            fcontrol = open(control, "w")
            for linea in lineas:
                #print(linea +"  "+ filePath1+'\n' )
                if (linea != filePath1+'\n' ):
                    fcontrol.write(linea)
                else:
                    pass
            fcontrol.close()
        else:
            print("archivo ocupado")
        
    
    def agregar_metadatos(self):
        
        metadato=str(self.text1.toPlainText())
        print(metadato)
        self.text1.setText("")
    

    
    #@QtCore.pyqtSlot()      
    def llenar_comboBox_marcas(self):
        self.metadatos1.clear()
        self.metadatos1.addItems(sorted(self.marcas.keys()))

    #@QtCore.pyqtSlot(str)    
    def llenar_comboBox_modelos(self,  marca):
        self.metadatos2.clear()
        self.metadatos2.addItems(self.marcas[marca])
      
   # @QtCore.pyqtSlot(str)    
    def llenar_comboBox_servidor2(self,  servidor2):
        self.ui.metadatos3.clear()
        self.ui.metadatos3.addItems(self.servidor2[servidor2])

       
                                                     

            


  
   
        
 
            
             
        
  
if __name__ == "__main__":
    app =  QtWidgets.QApplication(sys.argv)
    window = Interfaz()
    window.show()
    sys.exit(app.exec_())
